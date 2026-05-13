"""
Adds Sheets 5–7 to AC_Inverter_Market_Data.xlsx with data from
additional sources: IEA, Mordor Intelligence, Grand View Research,
Business Research Insights, Reanin.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = "/mnt/c/Users/PrietoGarciaManuel/Dashboard_development-Branch1/AC_Inverter_Market_Data.xlsx"
wb = openpyxl.load_workbook(PATH)

# ─── Colour palette ────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
TEAL        = "1F6B75"
PURPLE      = "4B2C6E"
LIGHT_BLUE  = "BDD7EE"
PALE_BLUE   = "DEEAF1"
PALE_GREEN  = "E2EFDA"
PALE_ORANGE = "FCE4D6"
GREEN       = "70AD47"
YELLOW      = "FFD966"
GREY_HEAD   = "D9D9D9"
WHITE       = "FFFFFF"
ALT1        = "F2F7FB"
ALT2        = WHITE

def hdr(cell, text, bg=DARK_BLUE, fg=WHITE, bold=True, wrap=True, size=10):
    cell.value = text
    cell.font  = Font(bold=bold, color=fg, size=size)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=wrap)

def val(cell, v, bold=False, color="000000", bg=None, align="left",
        wrap=False, size=10):
    cell.value = v
    cell.font  = Font(bold=bold, color=color, size=size)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)

def link(cell, url, text=None):
    cell.hyperlink = url
    cell.value     = text or url
    cell.font      = Font(color="0563C1", underline="single", size=10)

def thin_border(ws, row_start, row_end, col_start, col_end):
    thin = Side(style="thin", color="BBBBBB")
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(r, c).border = Border(left=thin, right=thin,
                                           top=thin, bottom=thin)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, text, col_span, bg=DARK_BLUE):
    ws.merge_cells(f"A1:{get_column_letter(col_span)}1")
    t = ws["A1"]
    t.value = text
    t.font  = Font(bold=True, color=WHITE, size=13)
    t.fill  = PatternFill("solid", fgColor=bg)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

def subtitle_row(ws, text, col_span, row=2):
    ws.merge_cells(f"A{row}:{get_column_letter(col_span)}{row}")
    s = ws.cell(row, 1)
    s.value = text
    s.font  = Font(italic=True, color="595959", size=9)
    s.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 16

# ══════════════════════════════════════════════════════════════════════════
# SHEET 5 – IEA Cooling Data
# ══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("IEA Cooling Data")
ws5.sheet_view.showGridLines = False
ws5.freeze_panes = "A4"

title_row(ws5, "IEA Space Cooling Data — AC Stock, Penetration & Efficiency", 7, bg=TEAL)
subtitle_row(ws5,
    "Sources: IEA Space Cooling (energy-system page); IEA Future of Cooling China; IEA Future of Cooling SE Asia",
    7)

hdrs5 = ["Category", "Metric / Indicator", "Value", "Year / Period",
         "Geography", "Notes", "Source & Link"]
for c, h in enumerate(hdrs5, 1):
    hdr(ws5.cell(3, c), h, bg=TEAL)
ws5.row_dimensions[3].height = 36

iea_rows = [
    # Category, Metric, Value, Year, Geography, Notes, Source_url, Link_text
    # --- IEA Space Cooling page ---
    ("Stock & Penetration", "Global residential AC units in operation", "1.5+ billion", "2022",
     "Global", "Tripled since 2000",
     "https://www.iea.org/energy-system/buildings/space-cooling", "IEA Space Cooling"),
    ("Stock & Penetration", "Share of global population owning AC", "37%", "2022",
     "Global", "Projected 45% by 2030",
     "https://www.iea.org/energy-system/buildings/space-cooling", "IEA Space Cooling"),
    ("Stock & Penetration", "AC penetration rate", "~5%", "2022",
     "Sub-Saharan Africa", "Very low access; major growth market",
     "https://www.iea.org/energy-system/buildings/space-cooling", "IEA Space Cooling"),
    ("Stock & Penetration", "AC penetration rate", "<20%", "2022",
     "India & Indonesia", "Rapidly growing with income rise",
     "https://www.iea.org/energy-system/buildings/space-cooling", "IEA Space Cooling"),
    ("Stock & Penetration", "AC penetration rate", "~30%", "2022",
     "Mexico & Brazil", "Significant fixed-speed share in entry segment",
     "https://www.iea.org/energy-system/buildings/space-cooling", "IEA Space Cooling"),
    ("Stock & Penetration", "AC penetration rate", ">85%", "2022",
     "Japan, South Korea, USA", "Saturated markets",
     "https://www.iea.org/energy-system/buildings/space-cooling", "IEA Space Cooling"),
    ("Energy & Emissions", "Space cooling share of buildings electricity", "Tripled since 1990", "2022",
     "Global", "Growing faster than any other end use",
     "https://www.iea.org/energy-system/buildings/space-cooling", "IEA Space Cooling"),
    ("Energy & Emissions", "Indirect CO₂ emissions from space cooling", "~1 Gt CO₂", "2022",
     "Global", "+2% vs 2021",
     "https://www.iea.org/energy-system/buildings/space-cooling", "IEA Space Cooling"),
    ("Energy & Emissions", "Target emissions reduction by 2030 (Net Zero)", "~40% of today's level", "2030 (NZE)",
     "Global", "Requires 3× faster efficiency improvement rate",
     "https://www.iea.org/energy-system/buildings/space-cooling", "IEA Space Cooling"),
    ("Policy Coverage", "Countries with AC MEPS", "90+", "2024",
     "Global", "86% of global residential consumption covered",
     "https://www.iea.org/energy-system/buildings/space-cooling", "IEA Space Cooling"),
    ("Policy Coverage", "Countries with AC energy labels", "95+", "2024",
     "Global", "Up from two-thirds in 2010",
     "https://www.iea.org/energy-system/buildings/space-cooling", "IEA Space Cooling"),
    # --- IEA Future of Cooling: China ---
    ("China – Production", "China share of global RAC production", "~70%", "2017",
     "China", "China is the dominant global manufacturer",
     "https://www.iea.org/reports/the-future-of-cooling-in-china", "IEA Future Cooling China"),
    ("China – Ownership", "Household AC ownership", "~60%", "2017",
     "China", "Projected ~85% by 2030; 1.1 billion units by 2030",
     "https://www.iea.org/reports/the-future-of-cooling-in-china", "IEA Future Cooling China"),
    ("China – Efficiency", "Variable-speed APF vs best available (2015-17)", "50–60% lower", "2015–2017",
     "China", "Average variable-speed was 20% below efficient models available in market",
     "https://www.iea.org/reports/the-future-of-cooling-in-china", "IEA Future Cooling China"),
    ("China – Efficiency", "Fixed-speed efficiency trend", "No improvement", "Since 2015",
     "China", "Sales average at or equal to least efficient available",
     "https://www.iea.org/reports/the-future-of-cooling-in-china", "IEA Future Cooling China"),
    ("China – Energy", "Cooling electricity use", "~400 TWh", "2017",
     "China", "8% of buildings sector; projected ~750 TWh by 2030 baseline",
     "https://www.iea.org/reports/the-future-of-cooling-in-china", "IEA Future Cooling China"),
    ("China – Energy", "Efficient scenario savings (2030)", "205 TWh (equipment) + 100 TWh (building)",
     "2030 (Efficient)", "China", "vs baseline; from variable-speed + building envelope",
     "https://www.iea.org/reports/the-future-of-cooling-in-china", "IEA Future Cooling China"),
    ("China – Capacity", "New residential mini-split/multi-split units 2017–2030", "380 million", "2017–2030",
     "China", "Plus 30 million ducted + 105 million non-residential",
     "https://www.iea.org/reports/the-future-of-cooling-in-china", "IEA Future Cooling China"),
    # --- IEA Future of Cooling: Southeast Asia ---
    ("SE Asia – Stock", "AC units in operation", "40 million", "2017",
     "Southeast Asia", "Only 15% household penetration",
     "https://www.iea.org/reports/the-future-of-cooling-in-southeast-asia", "IEA Future Cooling SE Asia"),
    ("SE Asia – Stock", "Projected AC units", "300 million", "2040",
     "Southeast Asia", "Half projected in Indonesia alone",
     "https://www.iea.org/reports/the-future-of-cooling-in-southeast-asia", "IEA Future Cooling SE Asia"),
    ("SE Asia – Penetration", "Household AC penetration", "15%", "2017",
     "Southeast Asia", "Very low baseline; rapid growth expected",
     "https://www.iea.org/reports/the-future-of-cooling-in-southeast-asia", "IEA Future Cooling SE Asia"),
    ("SE Asia – Efficiency", "Peak cooling demand share by 2040 (current policies)", "30%", "2040 (CPP)",
     "Southeast Asia", "Of region's peak electricity demand",
     "https://www.iea.org/reports/the-future-of-cooling-in-southeast-asia", "IEA Future Cooling SE Asia"),
    ("SE Asia – Efficiency", "Max locally manufactured efficiency – Thailand", "5.6 W/W", "2018",
     "Thailand", "Best available; ASEAN SHINE minimum = 3.08 W/W",
     "https://www.iea.org/reports/the-future-of-cooling-in-southeast-asia", "IEA Future Cooling SE Asia"),
    ("SE Asia – Efficiency", "Max locally manufactured efficiency – Vietnam", "5.6 W/W", "2018",
     "Vietnam", "Same as Thailand best-in-market",
     "https://www.iea.org/reports/the-future-of-cooling-in-southeast-asia", "IEA Future Cooling SE Asia"),
    ("SE Asia – Efficiency", "Minimum locally manufactured efficiency (all markets)", "3.7 W/W", "2018",
     "SE Asia (all)", "Floor across all analysed ASEAN markets",
     "https://www.iea.org/reports/the-future-of-cooling-in-southeast-asia", "IEA Future Cooling SE Asia"),
    ("SE Asia – Efficiency", "Target SEER by 2030", "6.0", "2030 target",
     "SE Asia", "Up from global average of 4.0",
     "https://www.iea.org/reports/the-future-of-cooling-in-southeast-asia", "IEA Future Cooling SE Asia"),
    ("SE Asia – Efficiency", "Target SEER by 2040", "8.0", "2040 target",
     "SE Asia", "Requires strong policy push",
     "https://www.iea.org/reports/the-future-of-cooling-in-southeast-asia", "IEA Future Cooling SE Asia"),
    ("SE Asia – Energy", "Energy savings potential by 2040 (stronger MEPS)", "Up to 110 TWh", "2040",
     "SE Asia", "CO₂ savings of ~30 Mt; NOx -75%, SO₂ -80%, PM2.5 -95%",
     "https://www.iea.org/reports/the-future-of-cooling-in-southeast-asia", "IEA Future Cooling SE Asia"),
]

alts = [ALT1, ALT2]
for i, r in enumerate(iea_rows, 4):
    cat, metric, value, year, geo, notes, url, link_text = r
    bg = alts[i % 2]
    # Color-code by category
    cat_colors = {
        "Stock & Penetration": "DEEAF1",
        "Energy & Emissions":  "FCE4D6",
        "Policy Coverage":     "E2EFDA",
        "China – Production":  "EAF0FB",
        "China – Ownership":   "EAF0FB",
        "China – Efficiency":  "EAF0FB",
        "China – Energy":      "EAF0FB",
        "China – Capacity":    "EAF0FB",
        "SE Asia – Stock":     "FFF2CC",
        "SE Asia – Penetration": "FFF2CC",
        "SE Asia – Efficiency":  "FFF2CC",
        "SE Asia – Energy":      "FFF2CC",
    }
    cat_bg = cat_colors.get(cat, bg)
    val(ws5.cell(i, 1), cat,    bold=True, bg=cat_bg, wrap=True, size=9)
    val(ws5.cell(i, 2), metric, wrap=True, bg=bg, size=9)
    val(ws5.cell(i, 3), value,  bold=True, align="center", bg=bg)
    val(ws5.cell(i, 4), year,   align="center", bg=bg, size=9)
    val(ws5.cell(i, 5), geo,    align="center", bg=bg, size=9)
    val(ws5.cell(i, 6), notes,  wrap=True, bg=bg, size=9)
    link(ws5.cell(i, 7), url, link_text)
    ws5.row_dimensions[i].height = 40

thin_border(ws5, 3, 3 + len(iea_rows), 1, 7)
set_col_widths(ws5, [22, 44, 20, 16, 18, 46, 26])

# ══════════════════════════════════════════════════════════════════════════
# SHEET 6 – Market Research Reports (Mordor / GVR / BRI / Reanin)
# ══════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Market Research Reports")
ws6.sheet_view.showGridLines = False
ws6.freeze_panes = "A4"

title_row(ws6,
    "Inverter AC Market Size, Shares & Forecasts — Commercial Research Reports",
    8, bg=PURPLE)
subtitle_row(ws6,
    "Sources: Mordor Intelligence; Grand View Research; Business Research Insights; Reanin. "
    "Note: These firms aggregate and model data — treat as indicative / cross-check only.",
    8)

hdrs6 = ["Metric / Indicator", "Value", "Year / Period",
         "Geography", "Segment", "Notes / Caveats",
         "Source", "Link"]
for c, h in enumerate(hdrs6, 1):
    hdr(ws6.cell(3, c), h, bg=PURPLE)
ws6.row_dimensions[3].height = 36

mkt_rows = [
    # ── Mordor Intelligence ──
    ("Total AC market size", "USD 171.88 billion", "2026",
     "Global", "All AC types", "Mordor Intelligence forecast",
     "Mordor Intelligence",
     "https://www.mordorintelligence.com/industry-reports/air-conditioner-market"),
    ("Total AC market size (projected)", "USD 239.01 billion", "2031",
     "Global", "All AC types", "CAGR 6.42% (2026–2031)",
     "Mordor Intelligence",
     "https://www.mordorintelligence.com/industry-reports/air-conditioner-market"),
    ("Inverter share of AC shipments", "71.85%", "2025",
     "Global", "Inverter segment", "Inverter CAGR 7.57% through 2031",
     "Mordor Intelligence",
     "https://www.mordorintelligence.com/industry-reports/air-conditioner-market"),
    ("Fixed/installed vs portable split", "90.48% fixed/installed", "2025",
     "Global", "Installation type", "Portable growing at 7.53% CAGR",
     "Mordor Intelligence",
     "https://www.mordorintelligence.com/industry-reports/air-conditioner-market"),
    ("Split systems share", "68.67%", "2025",
     "Global", "Product type", "Dominant product type globally",
     "Mordor Intelligence",
     "https://www.mordorintelligence.com/industry-reports/air-conditioner-market"),
    ("Residential share", "71.49%", "2025",
     "Global", "End-use segment", "Commercial growing at 7.21% CAGR",
     "Mordor Intelligence",
     "https://www.mordorintelligence.com/industry-reports/air-conditioner-market"),
    ("<3 kW systems share", "58.32%", "2025",
     "Global", "Capacity segment", "Typical residential unit range",
     "Mordor Intelligence",
     "https://www.mordorintelligence.com/industry-reports/air-conditioner-market"),
    ("B2C retail share", "74.84%", "2025",
     "Global", "Distribution channel", "",
     "Mordor Intelligence",
     "https://www.mordorintelligence.com/industry-reports/air-conditioner-market"),
    ("Asia-Pacific revenue share", "47.01%", "2025",
     "Asia-Pacific", "Regional", "CAGR 7.17%; ME&A fastest at 7.52%",
     "Mordor Intelligence",
     "https://www.mordorintelligence.com/industry-reports/air-conditioner-market"),
    # ── Grand View Research ──
    ("Total AC systems market size", "USD 98.74 billion", "2024",
     "Global", "All AC systems", "",
     "Grand View Research",
     "https://www.grandviewresearch.com/industry-analysis/air-conditioning-systems-industry"),
    ("Total AC systems market (projected)", "USD 182.97 billion", "2033",
     "Global", "All AC systems", "CAGR 7.2% (2025–2033)",
     "Grand View Research",
     "https://www.grandviewresearch.com/industry-analysis/air-conditioning-systems-industry"),
    ("Inverter AC market share", "68.6%", "2024",
     "Global", "Inverter segment", "Inverter CAGR 8.4% vs non-inverter 4.0% (2025–2033)",
     "Grand View Research",
     "https://www.grandviewresearch.com/industry-analysis/air-conditioning-systems-industry"),
    ("Ductless mini-split share", "49.4%", "2024",
     "Global", "Product type", "Largest product segment",
     "Grand View Research",
     "https://www.grandviewresearch.com/industry-analysis/air-conditioning-systems-industry"),
    ("Residential end-use share", "43.9%", "2024",
     "Global", "End-use", "Residential CAGR 8.9% (2025–2033)",
     "Grand View Research",
     "https://www.grandviewresearch.com/industry-analysis/air-conditioning-systems-industry"),
    ("Asia-Pacific revenue share", "57.4%", "2024",
     "Asia-Pacific", "Regional", "Largest region; India CAGR 9.5%",
     "Grand View Research",
     "https://www.grandviewresearch.com/industry-analysis/air-conditioning-systems-industry"),
    ("India CAGR", "9.5%", "2025–2033",
     "India", "Regional growth", "Fastest-growing major market",
     "Grand View Research",
     "https://www.grandviewresearch.com/industry-analysis/air-conditioning-systems-industry"),
    # ── Business Research Insights ──
    ("Inverter AC market size", "USD 18.43 billion", "2025",
     "Global", "Inverter-only segment", "Narrower segment definition than GVR/Mordor",
     "Business Research Insights",
     "https://www.businessresearchinsights.com/market-reports/inverter-technology-air-conditioner-market-124205"),
    ("Inverter AC market (projected)", "USD 33.55 billion", "2035",
     "Global", "Inverter-only segment", "CAGR 6.18–6.2% (2026–2035)",
     "Business Research Insights",
     "https://www.businessresearchinsights.com/market-reports/inverter-technology-air-conditioner-market-124205"),
    ("Asia-Pacific share of inverter AC market", "~70%", "2024",
     "Asia-Pacific", "Inverter segment", "Driven by China, India, Japan",
     "Business Research Insights",
     "https://www.businessresearchinsights.com/market-reports/inverter-technology-air-conditioner-market-124205"),
    ("Residential vs commercial split (inverter)", "65% residential / 35% commercial", "2024",
     "Global", "Inverter segment", "",
     "Business Research Insights",
     "https://www.businessresearchinsights.com/market-reports/inverter-technology-air-conditioner-market-124205"),
    ("Top 3 companies market concentration", "50%+ global share", "2024",
     "Global", "Market structure", "Daikin, LG Electronics, Mitsubishi Electric",
     "Business Research Insights",
     "https://www.businessresearchinsights.com/market-reports/inverter-technology-air-conditioner-market-124205"),
    ("IoT-enabled inverter AC adoption growth", "+40%", "2024",
     "Global", "Smart AC sub-segment", "Smart/IoT-connected inverter ACs",
     "Business Research Insights",
     "https://www.businessresearchinsights.com/market-reports/inverter-technology-air-conditioner-market-124205"),
    # ── Reanin ──
    ("Inverter AC market value", "USD 46,667 million", "2025",
     "Global", "Inverter-only segment", "Different methodology from BRI",
     "Reanin",
     "https://www.reanin.com/reports/global-inverter-technology-air-conditioner-market"),
    ("Inverter AC market (projected)", "USD 79,979 million", "2032",
     "Global", "Inverter-only segment", "CAGR 8.0% (2026–2032)",
     "Reanin",
     "https://www.reanin.com/reports/global-inverter-technology-air-conditioner-market"),
    ("Users opting for inverter AC", "65%", "2024",
     "Global", "Consumer adoption", "Share of buyers choosing inverter",
     "Reanin",
     "https://www.reanin.com/reports/global-inverter-technology-air-conditioner-market"),
    ("Residential AC sales with inverter", "60%", "2024",
     "Global", "Residential segment", "Of all new residential AC sales",
     "Reanin",
     "https://www.reanin.com/reports/global-inverter-technology-air-conditioner-market"),
    ("New AC models with inverter tech", "55%", "2024",
     "Global", "Product launches", "Of newly launched models",
     "Reanin",
     "https://www.reanin.com/reports/global-inverter-technology-air-conditioner-market"),
    ("Installations using inverter systems", "60%", "Recent years",
     "Global", "Installation", "Of all recent AC installations",
     "Reanin",
     "https://www.reanin.com/reports/global-inverter-technology-air-conditioner-market"),
    ("Countries with EE standards favouring inverter", "45+", "2024",
     "Global", "Policy", "Countries that have implemented EE standards for ACs",
     "Reanin",
     "https://www.reanin.com/reports/global-inverter-technology-air-conditioner-market"),
    ("Asia-Pacific installed inverter AC units", "~45% of global", "2024",
     "Asia-Pacific", "Regional share", ">50% overall AC market share",
     "Reanin",
     "https://www.reanin.com/reports/global-inverter-technology-air-conditioner-market"),
    ("Energy savings vs fixed-speed", "Up to 30% less electricity", "2024",
     "Global", "Efficiency", "Widely cited across all sources",
     "Reanin / IEA",
     "https://www.reanin.com/reports/global-inverter-technology-air-conditioner-market"),
]

src_colors = {
    "Mordor Intelligence":        "EAF0FB",
    "Grand View Research":        "E2EFDA",
    "Business Research Insights": "FFF2CC",
    "Reanin":                     "FCE4D6",
    "Reanin / IEA":               "FCE4D6",
}

for i, r in enumerate(mkt_rows, 4):
    metric, value, year, geo, seg, notes, src, url = r
    bg = src_colors.get(src, ALT1)
    val(ws6.cell(i, 1), metric, wrap=True, bg=bg, size=9)
    val(ws6.cell(i, 2), value,  bold=True, align="center", bg=bg)
    val(ws6.cell(i, 3), year,   align="center", bg=bg, size=9)
    val(ws6.cell(i, 4), geo,    align="center", bg=bg, size=9)
    val(ws6.cell(i, 5), seg,    align="center", bg=bg, size=9, wrap=True)
    val(ws6.cell(i, 6), notes,  wrap=True, bg=bg, size=9)
    val(ws6.cell(i, 7), src,    bold=True, bg=bg, size=9, wrap=True)
    link(ws6.cell(i, 8), url, "Source")
    ws6.row_dimensions[i].height = 36

# Legend / colour key
leg_row = 4 + len(mkt_rows) + 1
legend_items = [
    ("EAF0FB", "Mordor Intelligence"),
    ("E2EFDA", "Grand View Research"),
    ("FFF2CC", "Business Research Insights"),
    ("FCE4D6", "Reanin"),
]
ws6.cell(leg_row, 1).value = "Colour key:"
ws6.cell(leg_row, 1).font  = Font(bold=True, size=9)
for j, (clr, lbl) in enumerate(legend_items, 2):
    c = ws6.cell(leg_row, j)
    c.value = lbl
    c.fill  = PatternFill("solid", fgColor=clr)
    c.font  = Font(size=9)
    c.alignment = Alignment(horizontal="center")

# Caveat note
cav_row = leg_row + 1
ws6.merge_cells(f"A{cav_row}:H{cav_row}")
cv = ws6.cell(cav_row, 1)
cv.value = ("CAVEAT: Market research firms use different scope definitions, methodologies, "
            "and base-year assumptions. Figures are not directly comparable across reports. "
            "Use for directional context only. For authoritative unit-volume data, refer to "
            "Sheets 1–4 (JARN/IIF) and Sheet 5 (IEA).")
cv.font  = Font(italic=True, size=9, color="595959")
cv.alignment = Alignment(wrap_text=True)
ws6.row_dimensions[cav_row].height = 44

thin_border(ws6, 3, 3 + len(mkt_rows), 1, 8)
set_col_widths(ws6, [36, 22, 14, 16, 20, 38, 22, 12])

# ══════════════════════════════════════════════════════════════════════════
# Update Sources & Methodology sheet (Sheet 4)
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb["Sources & Methodology"]

# Find the last data row and add new sources after it
# Existing sources occupy rows 3–10 (8 sources); methodology note at row 11
# We'll append new entries starting at row 12
# Original sheet has 8 sources (rows 3-10) + note row at 12; start after that
next_row = 14

new_sources = [
    ("9", "IEA – Space Cooling (energy system page)",
     "International Energy Agency (IEA)",
     "Free / Open Access",
     "https://www.iea.org/energy-system/buildings/space-cooling"),
    ("10", "IEA – The Future of Cooling in China",
     "International Energy Agency (IEA)",
     "Free / Open Access (report summary)",
     "https://www.iea.org/reports/the-future-of-cooling-in-china"),
    ("11", "IEA – The Future of Cooling in Southeast Asia",
     "International Energy Agency (IEA)",
     "Free / Open Access (report summary)",
     "https://www.iea.org/reports/the-future-of-cooling-in-southeast-asia"),
    ("12", "Mordor Intelligence – Air Conditioner Market Report (2026–2031)",
     "Mordor Intelligence",
     "Paywalled (summary page free)",
     "https://www.mordorintelligence.com/industry-reports/air-conditioner-market"),
    ("13", "Grand View Research – Air Conditioning Systems Market (2025–2033)",
     "Grand View Research",
     "Paywalled (summary page free)",
     "https://www.grandviewresearch.com/industry-analysis/air-conditioning-systems-industry"),
    ("14", "Business Research Insights – Inverter Technology Air Conditioner Market (2026–2035)",
     "Business Research Insights",
     "Paywalled (summary page free)",
     "https://www.businessresearchinsights.com/market-reports/inverter-technology-air-conditioner-market-124205"),
    ("15", "Reanin – Global Inverter Technology Air Conditioner Market (2026–2032)",
     "Reanin",
     "Paywalled (summary page free)",
     "https://www.reanin.com/reports/global-inverter-technology-air-conditioner-market"),
]

thin = Side(style="thin", color="BBBBBB")
alts = [ALT1, ALT2]
PALE_ORANGE2 = "FCE4D6"

for i, (num, name, pub, access, url) in enumerate(new_sources):
    row = next_row + i
    bg = alts[row % 2]
    val(ws4.cell(row, 1), num, align="center", bg=bg, bold=True)
    val(ws4.cell(row, 2), name, wrap=True, bg=bg, size=9)
    val(ws4.cell(row, 3), pub,  wrap=True, bg=bg, size=9)
    val(ws4.cell(row, 4), access, wrap=True,
        bg=PALE_ORANGE2 if "Paywalled" in access or "PAID" in access else bg, size=9)
    link(ws4.cell(row, 5), url, url)
    ws4.row_dimensions[row].height = 36
    for c in range(1, 6):
        ws4.cell(row, c).border = Border(
            left=thin, right=thin, top=thin, bottom=thin)

# Save
wb.save(PATH)
print(f"Updated: {PATH}")
print(f"Sheets: {wb.sheetnames}")
