"""
build_dashboard_excel.py
Produces CoolProgress_Dashboard_Data_Review.xlsx for partner review.

Sheet 1 – Emissions Trajectory (global aggregates)
Sheet 2 – Emissions Trajectory (country detail)
Sheet 3 – Appliance Growth (global aggregates)
Sheet 4 – Appliance Growth (country detail)
Sheet 5 – Data Sources & Methodology
"""

import json
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent.parent

# ── colours ──────────────────────────────────────────────────────────────────
C_HEADER_DARK  = "1E3A2F"   # dark green
C_HEADER_MID   = "2D7D5A"   # mid green
C_HEADER_CLASP = "0369A1"   # CLASP blue
C_HEADER_GCI   = "7C3AED"   # GCI purple
C_ACCENT       = "F0FDF4"   # light green row tint
C_ACCENT2      = "EFF6FF"   # light blue row tint
C_WARN         = "FEF9C3"   # yellow highlight
WHITE          = "FFFFFF"
LIGHT_GREY     = "F8FAFC"

thin = Side(style="thin", color="CBD5E1")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hfill(hex_str):
    return PatternFill("solid", fgColor=hex_str)

def header_cell(ws, row, col, value, bg=C_HEADER_DARK, fg=WHITE, bold=True, size=10, wrap=True, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = hfill(bg)
    c.font = Font(bold=bold, color=fg, size=size)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = BORDER
    return c

def data_cell(ws, row, col, value, number_format=None, bg=None, bold=False, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = hfill(bg)
    c.font = Font(bold=bold, size=9)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = BORDER
    if number_format:
        c.number_format = number_format
    return c

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── load raw JSON ─────────────────────────────────────────────────────────────
clasp_traj  = json.loads((BASE / "clasp_trajectory_raw.json").read_text(encoding="utf-8"))
clasp_growth = json.loads((BASE / "clasp_growth_raw.json").read_text(encoding="utf-8"))
gci_raw      = json.loads((BASE / "gci_subcool_raw.json").read_text(encoding="utf-8"))

TRAJ_YEARS   = [2020, 2025, 2030, 2035, 2040, 2045, 2050]
GROWTH_YEARS = sorted({r["year"] for r in clasp_growth})
APPLIANCES   = ["Air Conditioning", "Ceiling and Portable Fans", "Refrigerator-Freezers"]
APP_SHORT    = {
    "Air Conditioning":       "Residential AC",
    "Ceiling and Portable Fans": "Ceiling Fans",
    "Refrigerator-Freezers":  "Refrigerators",
}
GCI_SCENARIOS = ["BAU", "KIP", "KIP_PLUS"]
GCI_SCN_LABEL = {"BAU": "BAU", "KIP": "Kigali Implementation", "KIP_PLUS": "Kigali+"}

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 – Emissions Trajectory: global aggregates
# ─────────────────────────────────────────────────────────────────────────────
def build_sheet1(wb):
    ws = wb.create_sheet("1. Emissions Trajectory (Global)")

    # aggregate CLASP
    # {year: {appliance: {bau, gb, bat}}}
    clasp_agg = defaultdict(lambda: defaultdict(lambda: {"bau":0,"gb":0,"bat":0}))
    for r in clasp_traj:
        y = r["year"]; app = r["appliance"]
        clasp_agg[y][app]["bau"] += r.get("bau_co2_mt") or 0
        clasp_agg[y][app]["gb"]  += r.get("gb_co2_mt")  or 0
        clasp_agg[y][app]["bat"] += r.get("bat_co2_mt") or 0

    # aggregate GCI (trajectory years only, BAU+KIP)
    gci_traj_raw = json.loads((BASE / "gci_subcool_raw.json").read_text(encoding="utf-8"))
    gci_agg = defaultdict(lambda: defaultdict(lambda: {"direct":0,"indirect":0}))
    for r in gci_traj_raw:
        if r["year"] not in TRAJ_YEARS: continue
        scn = r["scenario_name"]
        gci_agg[r["year"]][scn]["direct"]   += r.get("direct_emission_mt")   or 0
        gci_agg[r["year"]][scn]["indirect"] += r.get("indirect_emission_mt") or 0

    # ── title block ──
    ws.merge_cells("A1:R1")
    c = ws["A1"]
    c.value = "EMISSIONS TRAJECTORY — SCENARIO COMPARISON  |  Global Aggregates"
    c.font = Font(bold=True, size=12, color=WHITE)
    c.fill = hfill(C_HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:R2")
    c = ws["A2"]
    c.value = ("Chart displayed on CoolProgress Dashboard → Emissions Pillar → 'Emissions Trajectory — Scenario Comparison'.  "
               "Values are global sums in Mt CO₂. Years: 2020, 2025, 2030, 2035, 2040, 2045, 2050.")
    c.font = Font(italic=True, size=9, color="475569")
    c.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 30

    # ── CLASP section ──
    row = 4
    ws.merge_cells(f"A{row}:J{row}")
    c = ws.cell(row=row, column=1, value="CLASP MEPSY — Indirect (Energy-Related) Emissions  |  Table: clasp_energy_consumption")
    c.font = Font(bold=True, size=10, color=WHITE); c.fill = hfill(C_HEADER_CLASP)
    c.alignment = Alignment(horizontal="left", vertical="center"); ws.row_dimensions[row].height = 18

    row += 1
    headers_clasp = ["Appliance", "Year", "BAU (Mt CO₂)", "Green Buildings (Mt CO₂)", "Best Available Tech (Mt CO₂)",
                     "GB Savings vs BAU (Mt)", "BAT Savings vs BAU (Mt)", "GB Reduction %", "BAT Reduction %", "Notes"]
    for ci, h in enumerate(headers_clasp, 1):
        header_cell(ws, row, ci, h, bg=C_HEADER_CLASP)
    ws.row_dimensions[row].height = 32

    row += 1
    alt = False
    for app in APPLIANCES:
        for yi, y in enumerate(TRAJ_YEARS):
            bau = clasp_agg[y][app]["bau"]
            gb  = clasp_agg[y][app]["gb"]
            bat = clasp_agg[y][app]["bat"]
            gb_sav  = bau - gb  if bau else 0
            bat_sav = bau - bat if bau else 0
            gb_pct  = gb_sav  / bau * 100 if bau else 0
            bat_pct = bat_sav / bau * 100 if bau else 0
            bg = C_ACCENT2 if alt else WHITE
            note = "Baseline — scenarios identical" if y <= 2025 else ""
            data_cell(ws, row, 1, APP_SHORT[app], bg=bg, align="left")
            data_cell(ws, row, 2, y,   bg=bg, number_format="0")
            data_cell(ws, row, 3, round(bau,2),     bg=bg, number_format="0.00")
            data_cell(ws, row, 4, round(gb,2),      bg=bg, number_format="0.00")
            data_cell(ws, row, 5, round(bat,2),     bg=bg, number_format="0.00")
            data_cell(ws, row, 6, round(gb_sav,2),  bg=bg, number_format="0.00")
            data_cell(ws, row, 7, round(bat_sav,2), bg=bg, number_format="0.00")
            data_cell(ws, row, 8, round(gb_pct,1),  bg=bg, number_format="0.0\%")
            data_cell(ws, row, 9, round(bat_pct,1), bg=bg, number_format="0.0\%")
            data_cell(ws, row, 10, note, bg=bg, align="left")
            row += 1
        alt = not alt
        # subtotal per appliance
        for scn_col, scn_key in [(3,"bau"),(4,"gb"),(5,"bat")]:
            total = sum(clasp_agg[y][app][scn_key] for y in TRAJ_YEARS)
        # blank separator
        row += 1

    # CLASP global total rows
    ws.merge_cells(f"A{row}:J{row}")
    c = ws.cell(row=row, column=1, value="CLASP GLOBAL TOTALS (all 3 appliances combined)")
    c.font = Font(bold=True, size=9, color=WHITE); c.fill = hfill("334155")
    c.alignment = Alignment(horizontal="left", vertical="center"); row += 1

    header_cell(ws, row, 1, "Year", bg="334155")
    header_cell(ws, row, 2, "BAU Total (Mt CO₂)", bg="334155")
    header_cell(ws, row, 3, "Green Buildings Total (Mt CO₂)", bg="334155")
    header_cell(ws, row, 4, "Best Available Tech Total (Mt CO₂)", bg="334155")
    header_cell(ws, row, 5, "BAT Savings vs BAU (Mt)", bg="334155")
    header_cell(ws, row, 6, "BAT Reduction %", bg="334155")
    row += 1

    for y in TRAJ_YEARS:
        bau_t = sum(clasp_agg[y][a]["bau"] for a in APPLIANCES)
        gb_t  = sum(clasp_agg[y][a]["gb"]  for a in APPLIANCES)
        bat_t = sum(clasp_agg[y][a]["bat"] for a in APPLIANCES)
        bat_sav = bau_t - bat_t
        bat_pct = bat_sav / bau_t * 100 if bau_t else 0
        bg = C_ACCENT if y % 10 == 0 else WHITE
        data_cell(ws, row, 1, y,              bg=bg, bold=True, number_format="0")
        data_cell(ws, row, 2, round(bau_t,1), bg=bg, bold=True, number_format="0.0")
        data_cell(ws, row, 3, round(gb_t,1),  bg=bg, number_format="0.0")
        data_cell(ws, row, 4, round(bat_t,1), bg=bg, number_format="0.0")
        data_cell(ws, row, 5, round(bat_sav,1), bg=bg, number_format="0.0")
        data_cell(ws, row, 6, round(bat_pct,1), bg=bg, number_format="0.0\%")
        row += 1

    row += 2

    # ── GCI section ──
    ws.merge_cells(f"A{row}:I{row}")
    c = ws.cell(row=row, column=1, value="GCI / HEAT GLOBAL COOLING MODEL — Direct (Refrigerant) Emissions  |  Table: global_model_subcool")
    c.font = Font(bold=True, size=10, color=WHITE); c.fill = hfill(C_HEADER_GCI)
    c.alignment = Alignment(horizontal="left", vertical="center"); ws.row_dimensions[row].height = 18; row += 1

    headers_gci = ["Year", "BAU Direct (Mt CO₂)", "BAU Indirect (Mt CO₂)", "BAU Total (Mt CO₂)",
                   "KIP Direct (Mt CO₂)", "KIP Savings vs BAU (Mt)", "KIP Reduction %",
                   "KIP+ Direct (Mt CO₂)", "Notes"]
    for ci, h in enumerate(headers_gci, 1):
        header_cell(ws, row, ci, h, bg=C_HEADER_GCI); row += 1

    for y in TRAJ_YEARS:
        bau_d = gci_agg[y]["BAU"]["direct"]; bau_i = gci_agg[y]["BAU"]["indirect"]
        kip_d = gci_agg[y]["KIP"]["direct"]; kip_p = gci_agg[y]["KIP_PLUS"]["direct"]
        kip_sav = bau_d - kip_d; kip_pct = kip_sav / bau_d * 100 if bau_d else 0
        note = "Baseline — no divergence before 2025" if y <= 2025 else ""
        bg = C_ACCENT2 if y % 10 == 0 else WHITE
        data_cell(ws, row, 1, y,              bold=True, bg=bg, number_format="0")
        data_cell(ws, row, 2, round(bau_d,1), bg=bg, number_format="0.0")
        data_cell(ws, row, 3, round(bau_i,1), bg=bg, number_format="0.0")
        data_cell(ws, row, 4, round(bau_d+bau_i,1), bg=bg, number_format="0.0")
        data_cell(ws, row, 5, round(kip_d,1), bg=bg, number_format="0.0")
        data_cell(ws, row, 6, round(kip_sav,1), bg=bg, number_format="0.0")
        data_cell(ws, row, 7, round(kip_pct,1), bg=bg, number_format="0.0\%")
        data_cell(ws, row, 8, round(kip_p,1),   bg=bg, number_format="0.0")
        data_cell(ws, row, 9, note, bg=bg, align="left")
        row += 1

    set_col_widths(ws, [22,8,16,22,22,18,18,14,14,30])
    ws.freeze_panes = "A6"
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 – Emissions Trajectory: country detail
# ─────────────────────────────────────────────────────────────────────────────
def build_sheet2(wb):
    ws = wb.create_sheet("2. Emissions Traj. (Country)")

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "EMISSIONS TRAJECTORY — COUNTRY-LEVEL DETAIL  |  Source: clasp_energy_consumption (CLASP) + global_model_subcool (GCI)"
    c.font = Font(bold=True, size=11, color=WHITE); c.fill = hfill(C_HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:I2")
    ws["A2"].value = "CLASP values = indirect (energy-related) CO₂ in Mt. GCI values = direct (refrigerant leaks) CO₂ in Mt."
    ws["A2"].font = Font(italic=True, size=9, color="475569")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 20

    row = 4
    # CLASP country detail
    ws.merge_cells(f"A{row}:I{row}")
    c = ws.cell(row=row, column=1, value="CLASP — Country Detail  |  Table: clasp_energy_consumption  |  Columns: bau_co2_mt, gb_co2_mt, bat_co2_mt")
    c.font = Font(bold=True, size=9, color=WHITE); c.fill = hfill(C_HEADER_CLASP)
    c.alignment = Alignment(horizontal="left"); row += 1

    hdrs = ["Country Code","Country Name","Appliance","Year","BAU (Mt CO₂)","Green Buildings (Mt CO₂)","Best Available Tech (Mt CO₂)","GB Savings (Mt)","BAT Savings (Mt)"]
    for ci, h in enumerate(hdrs, 1):
        header_cell(ws, row, ci, h, bg=C_HEADER_CLASP)
    ws.row_dimensions[row].height = 28; row += 1

    alt = False
    for r in sorted(clasp_traj, key=lambda x: (x["country_name"] or "", x["appliance"], x["year"])):
        bau = r.get("bau_co2_mt") or 0; gb = r.get("gb_co2_mt") or 0; bat = r.get("bat_co2_mt") or 0
        bg = C_ACCENT2 if alt else WHITE
        data_cell(ws, row, 1, r["country_code"], bg=bg)
        data_cell(ws, row, 2, r["country_name"],  bg=bg, align="left")
        data_cell(ws, row, 3, APP_SHORT.get(r["appliance"], r["appliance"]), bg=bg, align="left")
        data_cell(ws, row, 4, r["year"],           bg=bg, number_format="0")
        data_cell(ws, row, 5, round(bau,4),        bg=bg, number_format="0.0000")
        data_cell(ws, row, 6, round(gb,4),         bg=bg, number_format="0.0000")
        data_cell(ws, row, 7, round(bat,4),        bg=bg, number_format="0.0000")
        data_cell(ws, row, 8, round(bau-gb,4),     bg=bg, number_format="0.0000")
        data_cell(ws, row, 9, round(bau-bat,4),    bg=bg, number_format="0.0000")
        row += 1; alt = not alt

    row += 2
    # GCI country detail
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1, value="GCI / HEAT — Country Detail  |  Table: global_model_subcool  |  Columns: direct_emission_mt, indirect_emission_mt")
    c.font = Font(bold=True, size=9, color=WHITE); c.fill = hfill(C_HEADER_GCI)
    c.alignment = Alignment(horizontal="left"); row += 1

    hdrs2 = ["Country Code","Country Name","Subsector","Year","Scenario","Direct (Mt CO₂)","Indirect (Mt CO₂)","Total (Mt CO₂)"]
    for ci, h in enumerate(hdrs2, 1):
        header_cell(ws, row, ci, h, bg=C_HEADER_GCI)
    ws.row_dimensions[row].height = 28; row += 1

    gci_traj_rows = [r for r in gci_raw if r["year"] in TRAJ_YEARS]
    alt = False
    for r in sorted(gci_traj_rows, key=lambda x: (x["country_name"] or "", x["subsector"], x["scenario_name"], x["year"])):
        d = r.get("direct_emission_mt") or 0; i = r.get("indirect_emission_mt") or 0
        bg = C_ACCENT if alt else WHITE
        data_cell(ws, row, 1, r["country_code"], bg=bg)
        data_cell(ws, row, 2, r["country_name"],  bg=bg, align="left")
        data_cell(ws, row, 3, r["subsector"],     bg=bg, align="left")
        data_cell(ws, row, 4, r["year"],           bg=bg, number_format="0")
        data_cell(ws, row, 5, GCI_SCN_LABEL.get(r["scenario_name"], r["scenario_name"]), bg=bg)
        data_cell(ws, row, 6, round(d,4),          bg=bg, number_format="0.0000")
        data_cell(ws, row, 7, round(i,4),          bg=bg, number_format="0.0000")
        data_cell(ws, row, 8, round(d+i,4),        bg=bg, number_format="0.0000")
        row += 1; alt = not alt

    set_col_widths(ws, [12,24,22,8,20,18,20,18,16])
    ws.freeze_panes = "A6"
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 – Appliance Growth: global aggregates
# ─────────────────────────────────────────────────────────────────────────────
def build_sheet3(wb):
    ws = wb.create_sheet("3. Appliance Growth (Global)")

    # aggregate CLASP growth data per year + appliance
    clasp_g = defaultdict(lambda: defaultdict(lambda: {"units":0,"energy":0,"indirect":0}))
    for r in clasp_growth:
        y = r["year"]; app = r["appliance"]
        clasp_g[y][app]["units"]    += r.get("appliance_units_in_use") or 0
        clasp_g[y][app]["energy"]   += r.get("bau_final_energy_twh") or 0
        clasp_g[y][app]["indirect"] += r.get("bau_co2_mt") or 0

    # aggregate GCI direct BAU per year + subsector
    gci_g = defaultdict(lambda: defaultdict(float))
    for r in gci_raw:
        if r["scenario_name"] != "BAU": continue
        gci_g[r["year"]][r["subsector"]] += r.get("direct_emission_mt") or 0

    GCI_SUBSECTOR_TO_APP = {
        "Split residential air conditioners": "Air Conditioning",
        "Domestic refrigeration": "Refrigerator-Freezers"
    }

    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = "GLOBAL APPLIANCE STOCK & EMISSIONS TRAJECTORY  |  Global Aggregates  (2025–2050)"
    c.font = Font(bold=True, size=12, color=WHITE); c.fill = hfill(C_HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:N2")
    ws["A2"].value = ("Stock (millions of units) and Energy (TWh) from clasp_energy_consumption (CLASP Mepsy). "
                      "Indirect CO₂ from clasp_energy_consumption (bau_co2_mt). "
                      "Direct CO₂ from global_model_subcool (direct_emission_mt, scenario=BAU). "
                      "Fans have no direct emissions (no refrigerant).")
    ws["A2"].font = Font(italic=True, size=9, color="475569")
    ws["A2"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 35

    row = 4
    for app in APPLIANCES:
        app_label = APP_SHORT[app]
        ws.merge_cells(f"A{row}:N{row}")
        c = ws.cell(row=row, column=1, value=f"{app_label.upper()}  |  Source: CLASP (stock, energy, indirect) + GCI (direct — AC & Refrigerators only)")
        c.font = Font(bold=True, size=10, color=WHITE)
        c.fill = hfill(C_HEADER_MID if app == "Air Conditioning" else C_HEADER_CLASP if app == "Refrigerator-Freezers" else "D97706")
        c.alignment = Alignment(horizontal="left", vertical="center"); ws.row_dimensions[row].height = 18; row += 1

        hdrs = ["Year",
                "Stock (M units)\nCLASP: appliance_units_in_use",
                "Energy Demand (TWh)\nCLASP: bau_final_energy_twh",
                "Indirect CO₂ (Mt)\nCLASP: bau_co2_mt",
                "Direct CO₂ (Mt)\nGCI: direct_emission_mt (BAU)",
                "Total CO₂ (Mt)\n(Indirect + Direct)",
                "Direct %\nof Total"]
        for ci, h in enumerate(hdrs, 1):
            header_cell(ws, row, ci, h, bg="334155")
        ws.row_dimensions[row].height = 40; row += 1

        # find matching GCI subsector
        gci_sub = None
        for sub, mapped_app in GCI_SUBSECTOR_TO_APP.items():
            if mapped_app == app:
                gci_sub = sub; break

        alt = False
        for y in GROWTH_YEARS:
            d = clasp_g[y][app]
            indirect = d["indirect"]
            direct = gci_g[y].get(gci_sub, 0) if gci_sub else 0
            total = indirect + direct
            pct = direct / total * 100 if total else 0
            bg = C_ACCENT if alt else WHITE
            data_cell(ws, row, 1, y,                    bold=True, bg=bg, number_format="0")
            data_cell(ws, row, 2, round(d["units"],2),  bg=bg, number_format="#,##0.00")
            data_cell(ws, row, 3, round(d["energy"],2), bg=bg, number_format="#,##0.00")
            data_cell(ws, row, 4, round(indirect,4),    bg=bg, number_format="0.0000")
            data_cell(ws, row, 5, round(direct,4) if gci_sub else "N/A (no refrigerant)", bg=bg, number_format="0.0000" if gci_sub else None)
            data_cell(ws, row, 6, round(total,4) if gci_sub else round(indirect,4), bg=bg, number_format="0.0000")
            data_cell(ws, row, 7, round(pct,1) if gci_sub else "0.0", bg=bg, number_format="0.0\%" if gci_sub else None)
            row += 1; alt = not alt
        row += 2

    set_col_widths(ws, [8, 22, 22, 20, 20, 18, 12])
    ws.freeze_panes = "A4"
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 4 – Appliance Growth: country detail
# ─────────────────────────────────────────────────────────────────────────────
def build_sheet4(wb):
    ws = wb.create_sheet("4. Appliance Growth (Country)")

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "APPLIANCE GROWTH — COUNTRY-LEVEL DETAIL  |  Source: clasp_energy_consumption (CLASP Mepsy)"
    c.font = Font(bold=True, size=11, color=WHITE); c.fill = hfill(C_HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:J2")
    ws["A2"].value = "Supabase table: clasp_energy_consumption | Columns: appliance_units_in_use, bau_final_energy_twh, bau_co2_mt | Years: 2025–2050"
    ws["A2"].font = Font(italic=True, size=9, color="475569")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    row = 4
    hdrs = ["Country Code","Country Name","Appliance","Year",
            "Stock (M units)\nCLASP: appliance_units_in_use",
            "Energy (TWh)\nCLASP: bau_final_energy_twh",
            "Indirect CO₂ (Mt)\nCLASP: bau_co2_mt"]
    for ci, h in enumerate(hdrs, 1):
        header_cell(ws, row, ci, h, bg=C_HEADER_CLASP)
    ws.row_dimensions[row].height = 40; row += 1

    alt = False
    for r in sorted(clasp_growth, key=lambda x: (x["country_name"] or "", x["appliance"], x["year"])):
        bg = C_ACCENT2 if alt else WHITE
        data_cell(ws, row, 1, r["country_code"], bg=bg)
        data_cell(ws, row, 2, r["country_name"],  bg=bg, align="left")
        data_cell(ws, row, 3, APP_SHORT.get(r["appliance"], r["appliance"]), bg=bg, align="left")
        data_cell(ws, row, 4, r["year"],          bg=bg, number_format="0")
        data_cell(ws, row, 5, round(r.get("appliance_units_in_use") or 0, 4), bg=bg, number_format="0.0000")
        data_cell(ws, row, 6, round(r.get("bau_final_energy_twh")  or 0, 4), bg=bg, number_format="0.0000")
        data_cell(ws, row, 7, round(r.get("bau_co2_mt")            or 0, 4), bg=bg, number_format="0.0000")
        row += 1; alt = not alt

    # GCI country detail appended below
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1, value="GCI / HEAT — Country Detail (BAU, 2025–2050)  |  Table: global_model_subcool  |  direct_emission_mt + indirect_emission_mt")
    c.font = Font(bold=True, size=9, color=WHITE); c.fill = hfill(C_HEADER_GCI)
    c.alignment = Alignment(horizontal="left"); row += 1

    hdrs2 = ["Country Code","Country Name","Subsector","Year","Scenario","Direct CO₂ (Mt)\ndirect_emission_mt","Indirect CO₂ (Mt)\nindirect_emission_mt","Total CO₂ (Mt)"]
    for ci, h in enumerate(hdrs2, 1):
        header_cell(ws, row, ci, h, bg=C_HEADER_GCI)
    ws.row_dimensions[row].height = 36; row += 1

    gci_bau_from25 = [r for r in gci_raw if r["scenario_name"] == "BAU" and r["year"] >= 2025]
    alt = False
    for r in sorted(gci_bau_from25, key=lambda x: (x["country_name"] or "", x["subsector"], x["year"])):
        d = r.get("direct_emission_mt") or 0; i = r.get("indirect_emission_mt") or 0
        bg = C_ACCENT if alt else WHITE
        data_cell(ws, row, 1, r["country_code"], bg=bg)
        data_cell(ws, row, 2, r["country_name"],  bg=bg, align="left")
        data_cell(ws, row, 3, r["subsector"],     bg=bg, align="left")
        data_cell(ws, row, 4, r["year"],           bg=bg, number_format="0")
        data_cell(ws, row, 5, "BAU",               bg=bg)
        data_cell(ws, row, 6, round(d,4),          bg=bg, number_format="0.0000")
        data_cell(ws, row, 7, round(i,4),          bg=bg, number_format="0.0000")
        data_cell(ws, row, 8, round(d+i,4),        bg=bg, number_format="0.0000")
        row += 1; alt = not alt

    set_col_widths(ws, [12,24,22,8,20,20,20,16])
    ws.freeze_panes = "A5"
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 5 – Data Sources & Methodology
# ─────────────────────────────────────────────────────────────────────────────
def build_sheet5(wb):
    ws = wb.create_sheet("5. Sources & Methodology")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 40

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "DATA SOURCES & METHODOLOGY — CoolProgress Dashboard | Emissions Pillar"
    c.font = Font(bold=True, size=13, color=WHITE); c.fill = hfill(C_HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 26

    rows_meta = [
        ("CLASP Mepsy Tool", "clasp_energy_consumption",
         "Indirect (energy-related) CO₂ emissions, appliance stock, energy demand",
         "bau_co2_mt, gb_co2_mt, bat_co2_mt, appliance_units_in_use, bau_final_energy_twh",
         "https://www.clasp.ngo/tools/mepsy/"),
        ("GCI / HEAT Global Cooling Model", "global_model_subcool",
         "Direct (refrigerant leak) CO₂ emissions and indirect. Scenarios: BAU, Kigali Implementation, Kigali+",
         "direct_emission_mt, indirect_emission_mt, scenario_name, subsector",
         "https://www.green-cooling-initiative.org/"),
    ]
    meta_hdrs = ["Data Partner","Supabase Table","What It Covers","Key Columns Used","URL"]
    row = 3
    for ci, h in enumerate(meta_hdrs, 1):
        header_cell(ws, row, ci, h, bg=C_HEADER_DARK)
    ws.row_dimensions[row].height = 22; row += 1

    for m in rows_meta:
        for ci, v in enumerate(m, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = Font(size=9); c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        ws.row_dimensions[row].height = 50; row += 1

    row += 1
    scenarios = [
        ("CLASP: BAU", "Business as Usual — no new policies beyond today's baseline. Column: bau_co2_mt"),
        ("CLASP: Green Buildings (GB)", "Building efficiency standards drive gradual efficiency improvements. Column: gb_co2_mt"),
        ("CLASP: Best Available Tech (BAT)", "Average efficiency of new appliances rises to match today's best available technology. Column: bat_co2_mt"),
        ("GCI: BAU", "Business as Usual — refrigerant use grows with stock. Scenario: BAU"),
        ("GCI: Kigali Implementation (KIP)", "Full Kigali Amendment HFC phase-down schedule implemented. Scenario: KIP"),
        ("GCI: Kigali+ (KIP_PLUS)", "Accelerated phase-down beyond Kigali Amendment targets. Scenario: KIP_PLUS"),
    ]
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value="SCENARIO DEFINITIONS")
    c.font = Font(bold=True, size=10, color=WHITE); c.fill = hfill("334155")
    c.alignment = Alignment(horizontal="left", vertical="center"); row += 1

    for ci, h in enumerate(["Scenario","Description"], 1):
        header_cell(ws, row, ci, h, bg="334155")
    row += 1
    for s, d in scenarios:
        ws.cell(row=row, column=1, value=s).font = Font(bold=True, size=9)
        ws.cell(row=row, column=2, value=d).font = Font(size=9)
        ws.cell(row=row, column=1).border = BORDER; ws.cell(row=row, column=2).border = BORDER
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 28; row += 1

    row += 1
    notes = [
        ("Unit note", "All CO₂ figures are in Mt CO₂e (megatonnes of CO₂ equivalent). Stock is in millions of units. Energy is in TWh."),
        ("Scope", "Residential AC, domestic refrigerators, and ceiling fans only. Commercial refrigeration and chillers are excluded."),
        ("Direct vs Indirect", "Indirect = CO₂ from electricity used to power appliances (CLASP). Direct = CO₂ from refrigerant leaks (GCI/HEAT)."),
        ("Fans", "Ceiling and portable fans have zero direct emissions (no refrigerant)."),
        ("Supabase endpoint", "https://hcpmdkkavtadgugrqohl.supabase.co/rest/v1/"),
        ("Extracted on", "2026-05-06"),
    ]
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value="ADDITIONAL NOTES")
    c.font = Font(bold=True, size=10, color=WHITE); c.fill = hfill("334155")
    c.alignment = Alignment(horizontal="left", vertical="center"); row += 1

    for label, note in notes:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=note)
        c1.font = Font(bold=True, size=9); c2.font = Font(size=9)
        c1.border = BORDER; c2.border = BORDER
        c1.alignment = Alignment(vertical="center"); c2.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 30; row += 1

    ws.merge_cells(f"B{row}:E{row}")
    ws.merge_cells(f"A{row}:A{row}")

    return ws


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

build_sheet1(wb)
build_sheet2(wb)
build_sheet3(wb)
build_sheet4(wb)
build_sheet5(wb)

out = BASE / "CoolProgress_Dashboard_Data_Review.xlsx"
wb.save(str(out))
print(f"Saved: {out}")
print(f"Sheets: {[s.title for s in wb.worksheets]}")
