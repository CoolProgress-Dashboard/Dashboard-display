"""
AC_Inverter_Share.xlsx
One focused table: inverter vs non-inverter share of AC sales,
per country/region and year, from all sources collected.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb  = openpyxl.Workbook()
ws  = wb.active
ws.title = "Inverter Share"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A5"

# ── palette ────────────────────────────────────────────────────────────────
DARK       = "1F3864"
MID        = "2E75B6"
WHITE      = "FFFFFF"
ROW_A      = "EBF3FB"
ROW_B      = "FFFFFF"
GREEN_DARK = "375623"
GREEN_MED  = "70AD47"
GREEN_LT   = "C6EFCE"
AMBER      = "FFEB9C"
RED_LT     = "FFC7CE"
GREY_LT    = "F2F2F2"
ORANGE_LT  = "FCE4D6"

# ── helpers ────────────────────────────────────────────────────────────────
def thin():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(cell, txt, bg=DARK, fg=WHITE, sz=10, wrap=True):
    cell.value = txt
    cell.font  = Font(bold=True, color=fg, size=sz)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=wrap)
    cell.border = thin()

def v(cell, val, bold=False, col="000000", bg=None, ha="left",
      wrap=False, sz=10, border=True):
    cell.value = val
    cell.font  = Font(bold=bold, color=col, size=sz)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=ha, vertical="center",
                                wrap_text=wrap)
    if border:
        cell.border = thin()

def lnk(cell, url, txt, sz=9):
    cell.hyperlink = url
    cell.value     = txt
    cell.font      = Font(color="0563C1", underline="single", size=sz)
    cell.border    = thin()

def inv_colour(pct_str):
    """Return fill colour based on inverter share value."""
    try:
        n = float(str(pct_str).replace("~","").replace("%","")
                               .replace(">","").replace("<","")
                               .strip().split("-")[0].split("–")[0])
        if n >= 75: return GREEN_LT
        if n >= 50: return AMBER
        return RED_LT
    except Exception:
        return GREY_LT

def conf_colour(c):
    return {"High": GREEN_LT, "Medium": AMBER, "Low": ORANGE_LT,
            "Low-Medium": ORANGE_LT, "Forecast": GREY_LT}.get(c, GREY_LT)

# ── title ──────────────────────────────────────────────────────────────────
ws.merge_cells("A1:K1")
t = ws["A1"]
t.value = "Inverter vs Non-Inverter Air Conditioner Share — by Country / Region and Year"
t.font  = Font(bold=True, color=WHITE, size=14)
t.fill  = PatternFill("solid", fgColor=DARK)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:K2")
ws["A2"].value = (
    "All data points extracted from: JARN/IIF (2019 & 2025 special issues), "
    "CLASP, ZERO US Partners, Mordor Intelligence, Grand View Research, Reanin, IEA. "
    "Share = % of new AC units sold / shipped."
)
ws["A2"].font      = Font(italic=True, color="595959", size=9)
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 16

# colour legend row 3
ws.merge_cells("A3:B3"); ws["A3"].value = "Colour key:"
ws["A3"].font = Font(bold=True, size=9)
ws["A3"].alignment = Alignment(horizontal="right", vertical="center")
legend = [
    ("C3", GREEN_LT, "≥ 75% inverter"),
    ("D3", AMBER,    "50–74% inverter"),
    ("E3", RED_LT,   "< 50% inverter"),
    ("F3", GREY_LT,  "Forecast / unclear"),
]
for cell_addr, colour, label in legend:
    c = ws[cell_addr]
    c.value = label
    c.fill  = PatternFill("solid", fgColor=colour)
    c.font  = Font(size=9)
    c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[3].height = 16

# ── column headers ─────────────────────────────────────────────────────────
cols = [
    "Country / Region", "Year", "Inverter\nShare (%)",
    "Non-Inverter\nShare (%)", "Segment /\nScope",
    "Data\nConfidence", "Source", "Source URL", "Notes"
]
for c, h in enumerate(cols, 1):
    hdr(ws.cell(4, c), h, bg=MID)
ws.row_dimensions[4].height = 44

# ══════════════════════════════════════════════════════════════════════════
# DATA — every inverter-share data point collected
# (Country, Year, Inv%, NonInv%, Segment, Confidence, Source_label, URL, Notes)
# ══════════════════════════════════════════════════════════════════════════
DATA = [
    # ── Japan ──────────────────────────────────────────────────────────────
    ("Japan",        2022, "~100%", "~0%",   "All RAC",
     "High",
     "CLASP / U4E (JARN secondary)",
     "https://www.clasp.ngo/research/all/chinas-meps-lead-to-major-ac-market-transformation/",
     "Top Runner Programme; essentially 100% for many years. "
     "2.7 AC units per household."),

    # ── South Korea ────────────────────────────────────────────────────────
    ("South Korea",  2022, "~85%",  "~15%",  "All RAC",
     "Medium",
     "CLASP / U4E (JARN secondary)",
     "https://www.clasp.ngo/research/all/chinas-meps-lead-to-major-ac-market-transformation/",
     "High penetration driven by strong EE labelling & MEPS."),

    # ── China ──────────────────────────────────────────────────────────────
    ("China",        2018, "70%",   "30%",   "Retail RAC",
     "High",
     "IIF/JARN – World AC Market Trends (Jan 2019 special issue)",
     "https://iifiir.org/en/news/world-air-conditioning-market-figures-and-technical-trends",
     "Grade-1 EE ACs = 44% of market. EE growth +14% YoY. "
     "Total RAC market: 59.5 million units (−2.3% vs 2017)."),

    ("China",        2020, "~60%",  "~40%",  "All RAC",
     "High",
     "CLASP – China MEPS Market Transformation",
     "https://www.clasp.ngo/research/all/chinas-meps-lead-to-major-ac-market-transformation/",
     "Pre-MEPS enforcement baseline. After 2022 enforcement: rapid shift."),

    ("China",        2021, "73.5%", "26.5%", "All RAC",
     "High",
     "ZERO US Partners (JARN 2021 data)",
     "https://zeroac.us/world-air-conditioner-market/",
     "Market grew +2.6% over 2020. AC penetration 117.7 per 100 households."),

    ("China",        2022, "~98%",  "~2%",   "All RAC",
     "High",
     "CLASP – China MEPS Market Transformation",
     "https://www.clasp.ngo/research/all/chinas-meps-lead-to-major-ac-market-transformation/",
     "Post GB 21455-2019 enforcement (July 2022). Fixed-speed effectively "
     "eliminated. Grade-1 share: 19% → 56% in 2 years. "
     "CO₂ prevented 2020-2030: min. 470 Mt."),

    # ── Thailand ───────────────────────────────────────────────────────────
    ("Thailand",     2016, "~30%",  "~70%",  "All RAC",
     "Medium",
     "IIF/JARN – World AC Market Trends (Jan 2019 special issue)",
     "https://iifiir.org/en/news/world-air-conditioning-market-figures-and-technical-trends",
     "Baseline year cited in JARN as 'two years prior' to 2018."),

    ("Thailand",     2018, "50%",   "50%",   "All RAC",
     "High",
     "IIF/JARN – World AC Market Trends (Jan 2019 special issue)",
     "https://iifiir.org/en/news/world-air-conditioning-market-figures-and-technical-trends",
     "Inverter share doubled from ~30% in ~2 years."),

    ("Thailand",     2019, "~60%",  "~40%",  "All RAC",
     "Forecast",
     "IIF/JARN – World AC Market Trends (Jan 2019 special issue)",
     "https://iifiir.org/en/news/world-air-conditioning-market-figures-and-technical-trends",
     "JARN forecast published in early 2019."),

    ("Thailand",     2020, "~80%",  "~20%",  "All RAC",
     "Forecast",
     "IIF/JARN – World AC Market Trends (Jan 2019 special issue)",
     "https://iifiir.org/en/news/world-air-conditioning-market-figures-and-technical-trends",
     "JARN forecast published in early 2019. "
     "Driven by TISI energy label + consumer cost awareness."),

    # ── India ──────────────────────────────────────────────────────────────
    ("India",        2017, "~15%",  "~85%",  "Split RAC",
     "High",
     "IIF/JARN – World AC Market Trends (Jan 2019 special issue)",
     "https://iifiir.org/en/news/world-air-conditioning-market-figures-and-technical-trends",
     "Inverter penetration nascent; 20-25% price premium a key barrier."),

    ("India",        2018, "35%",   "65%",   "Split RAC",
     "High",
     "IIF/JARN – World AC Market Trends (Jan 2019 special issue)",
     "https://iifiir.org/en/news/world-air-conditioning-market-figures-and-technical-trends",
     "Share more than doubled YoY. BEE Star Label driving adoption."),

    # ── Malaysia ───────────────────────────────────────────────────────────
    ("Malaysia",     2020, "~65%",  "~35%",  "Retail shelf\n(survey)",
     "Medium",
     "CLASP / U4E (JARN secondary)",
     "https://www.clasp.ngo/research/all/chinas-meps-lead-to-major-ac-market-transformation/",
     "Based on share of inverter models available in surveyed retail stores "
     "(not sales volume). SIRIM EE label drives adoption."),

    # ── USA ────────────────────────────────────────────────────────────────
    ("USA – Unitary\n(central / ducted)", 2019, "~4%", "~96%", "Unitary segment",
     "High",
     "IIF/JARN – World AC Market Trends (Jan 2019 special issue)",
     "https://iifiir.org/en/news/world-air-conditioning-market-figures-and-technical-trends",
     "Inverter very rare in traditional US ducted/central AC market. "
     "SEER targets 20-24 but achieved via fixed-speed coil improvements. "
     "Cost premium 20-25% cited as barrier. "
     "Mini-split segment has much higher inverter share."),

    ("USA – Mini-split", 2024, "Majority", "Minority", "Mini-split RAC",
     "Low-Medium",
     "IIF/JARN – Global AC Market 2024 (Jan 2025 special issue summary)",
     "https://iifiir.org/en/news/global-air-conditioner-market-2024-performance",
     "Mini-split market declined −3.7% in 2024. "
     "Mini VRF (all inverter) grew double-digit. "
     "No exact % given in public sources."),

    # ── SE Asia (overall) ──────────────────────────────────────────────────
    ("Southeast Asia\n(overall)", 2024, "Growing\nrapidly", "Declining",
     "All RAC",
     "Low-Medium",
     "IIF/JARN – Global AC Market 2024 (Jan 2025 special issue summary)",
     "https://iifiir.org/en/news/global-air-conditioner-market-2024-performance",
     "Daikin reported +30% surge in inverter sales across SE Asia in 2024. "
     "Total demand +13.9%. No exact % in public sources. "
     "Household penetration only 15% (IEA 2017); rapid growth underway."),

    # ── Africa ─────────────────────────────────────────────────────────────
    ("Africa", 2021, "<20%", ">80%", "All RAC",
     "Low",
     "ZERO US Partners (JARN 2021 data)",
     "https://zeroac.us/world-air-conditioner-market/",
     "3.44 million units, +3.1%. Import taxes >40% suppress inverter "
     "adoption by amplifying the price premium. No exact % in public sources."),

    ("Africa\n(North Africa)", 2024, "<30%", ">70%", "All RAC",
     "Low",
     "IIF/JARN – Global AC Market 2024 (Jan 2025 special issue summary)",
     "https://iifiir.org/en/news/global-air-conditioner-market-2024-performance",
     "North Africa drove growth: 3.66 million units (+4.3%). "
     "Sub-Saharan Africa household penetration ~5% (IEA)."),

    # ── Middle East ────────────────────────────────────────────────────────
    ("Middle East",  2024, "Moderate", "Moderate", "All AC",
     "Low",
     "IIF/JARN – Global AC Market 2024 (Jan 2025 special issue summary)",
     "https://iifiir.org/en/news/global-air-conditioner-market-2024-performance",
     "Total AC market 10.85 million units. Gulf states higher inverter share "
     "due to intense cooling demand and recent energy subsidy reforms (GCC/ESMA/SASO)."),

    # ── Latin America ──────────────────────────────────────────────────────
    ("Latin America", 2021, "Increasing", "Dominant", "All RAC",
     "Low",
     "ZERO US Partners (JARN 2021 data)",
     "https://zeroac.us/world-air-conditioner-market/",
     "9.37 million units, +10.3%. Brazil/Mexico/Argentina = 77% of region. "
     "Brazil PROCEL label most advanced in region."),

    # ── Europe ─────────────────────────────────────────────────────────────
    ("Europe",       2021, "Majority", "Minority", "RAC / heat pumps",
     "Medium",
     "ZERO US Partners (JARN 2021 data)",
     "https://zeroac.us/world-air-conditioner-market/",
     "7.45 million units, +4.9%. EU Ecodesign/ErP directive effectively "
     "mandates inverter in most product classes."),

    ("Europe",       2024, "Majority", "Minority", "RAC / heat pumps",
     "Medium",
     "IIF/JARN – Global AC Market 2024 (Jan 2025 special issue summary)",
     "https://iifiir.org/en/news/global-air-conditioner-market-2024-performance",
     "Sluggish overall (cooler summer); southern Europe grew. "
     "EU Ecodesign + REPowerEU heat pump push maintains inverter dominance."),

    # ── Global ─────────────────────────────────────────────────────────────
    ("Global",       2018, "~40–45%", "~55–60%", "All RAC",
     "Low",
     "IIF/JARN – World AC Market Trends (Jan 2019 special issue)",
     "https://iifiir.org/en/news/world-air-conditioning-market-figures-and-technical-trends",
     "Total global RAC market 130.1 million units (+0.6% YoY). "
     "Global figure estimated; not stated explicitly in source."),

    ("Global",       2024, "~68–70%", "~30–32%", "All RAC",
     "Medium",
     "Grand View Research / IIF/JARN Jan 2025 summary",
     "https://www.grandviewresearch.com/industry-analysis/air-conditioning-systems-industry",
     "Total RAC market 140.58 million units (+4% YoY). "
     "Grand View Research: 68.6% inverter share (2024). "
     "Cross-checked with Mordor Intelligence (71.85% in 2025) and Reanin (60–65%)."),

    ("Global",       2025, "71.85%", "~28%", "All AC (shipments)",
     "Medium",
     "Mordor Intelligence – Air Conditioner Market Report",
     "https://www.mordorintelligence.com/industry-reports/air-conditioner-market",
     "Mordor Intelligence 2025 base-year figure. "
     "Inverter CAGR 7.57% vs non-inverter through 2031."),
]

# ── write rows ─────────────────────────────────────────────────────────────
alts = [ROW_A, ROW_B]
for i, row in enumerate(DATA, 5):
    country, year, inv, non_inv, seg, conf, src, url, notes = row
    bg   = alts[i % 2]
    ibg  = inv_colour(inv)
    cbg  = conf_colour(conf)

    v(ws.cell(i,1), country, bold=True,  bg=bg,  sz=10, wrap=True)
    v(ws.cell(i,2), year,    ha="center",bg=bg,  sz=10)
    v(ws.cell(i,3), inv,     bold=True,  bg=ibg, ha="center", sz=11)
    v(ws.cell(i,4), non_inv, bg=bg,      ha="center", sz=10)
    v(ws.cell(i,5), seg,     bg=bg,      ha="center", sz=9, wrap=True)
    v(ws.cell(i,6), conf,    bg=cbg,     ha="center", sz=9, bold=True)
    v(ws.cell(i,7), src,     bg=bg,      sz=8,  wrap=True)
    lnk(ws.cell(i,8), url, "→ source")
    v(ws.cell(i,9), notes,   bg=bg,      sz=8,  wrap=True)
    ws.row_dimensions[i].height = 50

# ── column widths ──────────────────────────────────────────────────────────
widths = [22, 8, 13, 13, 16, 12, 38, 12, 62]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── freeze & auto-filter ───────────────────────────────────────────────────
ws.auto_filter.ref = f"A4:{get_column_letter(len(cols))}{4 + len(DATA)}"

# ── source summary tab ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("Data Sources")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:D1")
ws2["A1"].value = "Sources used for Inverter Share data"
ws2["A1"].font  = Font(bold=True, color=WHITE, size=12)
ws2["A1"].fill  = PatternFill("solid", fgColor=DARK)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

for c, h in enumerate(["Source", "Publisher", "Access", "URL"], 1):
    hdr(ws2.cell(2, c), h, bg=MID)
ws2.row_dimensions[2].height = 28

sources = [
    ("IIF/JARN – World AC Market: Figures & Technical Trends\n(Jan 2019 special issue summary)",
     "IIF / JARN", "Free (IIF summary; full JARN paywalled)",
     "https://iifiir.org/en/news/world-air-conditioning-market-figures-and-technical-trends"),
    ("IIF/JARN – Global AC Market: 2024 Performance\n(Jan 2025 special issue summary)",
     "IIF / JARN", "Free (IIF summary; full JARN paywalled)",
     "https://iifiir.org/en/news/global-air-conditioner-market-2024-performance"),
    ("CLASP – China's MEPS Lead to Major AC Market Transformation",
     "CLASP", "Free / Open Access",
     "https://www.clasp.ngo/research/all/chinas-meps-lead-to-major-ac-market-transformation/"),
    ("ZERO US Partners – World Air Conditioner Market (JARN 2021 data)",
     "ZERO US Partners", "Free",
     "https://zeroac.us/world-air-conditioner-market/"),
    ("Mordor Intelligence – Air Conditioner Market Report (2026–2031)",
     "Mordor Intelligence", "Summary free; full report paywalled",
     "https://www.mordorintelligence.com/industry-reports/air-conditioner-market"),
    ("Grand View Research – Air Conditioning Systems Market (2025–2033)",
     "Grand View Research", "Summary free; full report paywalled",
     "https://www.grandviewresearch.com/industry-analysis/air-conditioning-systems-industry"),
    ("IEA – Space Cooling (energy system page)",
     "IEA", "Free",
     "https://www.iea.org/energy-system/buildings/space-cooling"),
    ("IEA – The Future of Cooling in Southeast Asia",
     "IEA", "Free (report summary)",
     "https://www.iea.org/reports/the-future-of-cooling-in-southeast-asia"),
    ("JARN – World AC Market 2025 Update (article 87132)",
     "JARN / eJARN", "PAYWALLED (subscription required)",
     "https://www.ejarn.com/article/detail/87132"),
]

alts2 = [ROW_A, ROW_B]
for i, (src, pub, acc, url) in enumerate(sources, 3):
    bg = alts2[i % 2]
    for c, val_txt in enumerate([src, pub, acc], 1):
        cell = ws2.cell(i, c)
        cell.value = val_txt
        cell.font  = Font(size=9,
                          color="CC0000" if "PAYWALLED" in val_txt else "000000")
        cell.fill  = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin()
    lnk(ws2.cell(i, 4), url, url)
    ws2.row_dimensions[i].height = 40

ws2.column_dimensions["A"].width = 52
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 28
ws2.column_dimensions["D"].width = 80

# ── save ───────────────────────────────────────────────────────────────────
out = "/mnt/c/Users/PrietoGarciaManuel/Dashboard_development-Branch1/AC_Inverter_Share.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Rows of data: {len(DATA)}")
